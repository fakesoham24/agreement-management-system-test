"""
Export Routes — Generate Excel exports of payment and agreement data with filter support.
"""
import io
from datetime import datetime
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from backend.auth import get_current_user
from backend.database import get_db

router = APIRouter(prefix="/api/export", tags=["Export"])


@router.get("/payments")
def export_payments_excel(
    year: int = Query(..., description="Year to filter payments"),
    month: int = Query(None, description="Month (0-indexed, 0=Jan) — omit for full year"),
    status: str = Query("", description="Filter by status: paid, pending, or empty for all"),
    current_user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    """Export filtered payments to an Excel (.xlsx) file. Admin only."""
    # Admin-only guard
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Export is restricted to administrators")

    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    cursor = db.cursor()

    # Build query based on role
    if current_user["role"] == "admin" or current_user.get("global_payment_access"):
        query = """
            SELECT p.id as payment_id, p.agreement_id, p.due_date, p.amount, p.status, p.paid_at,
                   aa.company_name, aa.currency, aa.payment_plans
            FROM payments p
            JOIN agreements a ON p.agreement_id = a.id
            LEFT JOIN agreement_analysis aa ON a.id = aa.agreement_id
            ORDER BY p.due_date ASC
        """
        params = []
    else:
        query = """
            SELECT p.id as payment_id, p.agreement_id, p.due_date, p.amount, p.status, p.paid_at,
                   aa.company_name, aa.currency, aa.payment_plans
            FROM payments p
            JOIN agreements a ON p.agreement_id = a.id
            LEFT JOIN agreement_analysis aa ON a.id = aa.agreement_id
            WHERE a.user_id = ?
            ORDER BY p.due_date ASC
        """
        params = [current_user["id"]]

    rows = cursor.execute(query, params).fetchall()

    # Process and filter payments (same logic as frontend)
    MONTH_FULL = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ]

    filtered = []
    for row in rows:
        r = dict(row)
        due_date = r.get("due_date")
        if not due_date:
            continue
        try:
            d = datetime.strptime(due_date, "%Y-%m-%d")
        except (ValueError, TypeError):
            continue

        if d.year != year:
            continue
        if month is not None and d.month - 1 != month:
            continue

        pay_status = (r.get("status") or "pending").lower()
        if status == "paid" and pay_status != "paid":
            continue
        if status == "pending" and pay_status == "paid":
            continue

        # Enrich with plan details
        plan_name = ""
        gst = 0
        tds = 0
        net = r["amount"]
        plans_str = r.get("payment_plans")
        if plans_str:
            try:
                plans = json.loads(plans_str) if isinstance(plans_str, str) else plans_str
                if isinstance(plans, list):
                    for plan in plans:
                        if isinstance(plan, dict):
                            plan_due = plan.get("due_date", "")
                            plan_amount = plan.get("amount", 0)
                            if plan_due == due_date and abs((plan_amount or 0) - (r["amount"] or 0)) < 1:
                                plan_name = plan.get("plan", "")
                                gst = plan.get("gst", 0) or 0
                                tds = plan.get("tds", 0) or 0
                                net = plan.get("net", 0) or r["amount"]
                                break
            except (json.JSONDecodeError, TypeError):
                pass

        filtered.append({
            "company_name": r.get("company_name") or "Unknown",
            "agreement_id": r["agreement_id"],
            "plan_name": plan_name or "—",
            "amount": r["amount"] or 0,
            "gst": gst,
            "tds": tds,
            "net": net,
            "due_date": due_date,
            "status": pay_status.capitalize(),
            "currency": r.get("currency") or "₹",
        })

    # Build Excel workbook
    wb = Workbook()
    ws = wb.active

    # Sheet title
    if month is not None and 0 <= month <= 11:
        period = f"{MONTH_FULL[month]} {year}"
    else:
        period = str(year)
    ws.title = f"Payments — {period}"[:31]  # Excel sheet name limit

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    total_font = Font(name="Calibri", bold=True, size=11)
    total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    paid_font = Font(name="Calibri", color="059669")
    pending_font = Font(name="Calibri", color="D97706")
    currency_format = '#,##0.00'

    # Title row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"Payment Report — {period}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle row
    ws.merge_cells("A2:I2")
    sub_cell = ws["A2"]
    status_label = f" | Status: {status.capitalize()}" if status else ""
    sub_cell.value = f"Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')}{status_label}"
    sub_cell.font = Font(name="Calibri", size=9, color="6B7280")
    ws.row_dimensions[2].height = 20

    # Headers (row 4)
    headers = ["Company", "Agreement ID", "Plan", "Amount", "GST (18%)", "TDS (10%)", "Net", "Due Date", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[4].height = 28

    # Data rows
    total_amount = 0
    total_gst = 0
    total_tds = 0
    total_net = 0

    for row_idx, p in enumerate(filtered, 5):
        currency = p["currency"]
        ws.cell(row=row_idx, column=1, value=p["company_name"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=f"AG-{str(p['agreement_id']).zfill(4)}").border = thin_border
        ws.cell(row=row_idx, column=3, value=p["plan_name"]).border = thin_border

        amount_cell = ws.cell(row=row_idx, column=4, value=p["amount"])
        amount_cell.number_format = currency_format
        amount_cell.border = thin_border

        gst_cell = ws.cell(row=row_idx, column=5, value=p["gst"])
        gst_cell.number_format = currency_format
        gst_cell.border = thin_border

        tds_cell = ws.cell(row=row_idx, column=6, value=p["tds"])
        tds_cell.number_format = currency_format
        tds_cell.border = thin_border

        net_cell = ws.cell(row=row_idx, column=7, value=p["net"])
        net_cell.number_format = currency_format
        net_cell.font = Font(name="Calibri", bold=True)
        net_cell.border = thin_border

        try:
            date_val = datetime.strptime(p["due_date"], "%Y-%m-%d")
            date_cell = ws.cell(row=row_idx, column=8, value=date_val)
            date_cell.number_format = "DD MMM YYYY"
        except (ValueError, TypeError):
            date_cell = ws.cell(row=row_idx, column=8, value=p["due_date"])
        date_cell.border = thin_border

        status_cell = ws.cell(row=row_idx, column=9, value=p["status"])
        status_cell.border = thin_border
        if p["status"] == "Paid":
            status_cell.font = paid_font
        else:
            status_cell.font = pending_font

        total_amount += p["amount"]
        total_gst += p["gst"]
        total_tds += p["tds"]
        total_net += p["net"]

    # Total row
    total_row = len(filtered) + 5
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=1).border = thin_border
    for col in range(2, 4):
        ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=col).border = thin_border

    for col, val in [(4, total_amount), (5, total_gst), (6, total_tds), (7, total_net)]:
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.number_format = currency_format
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border

    for col in [8, 9]:
        ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=col).border = thin_border

    # Auto-size columns
    col_widths = [28, 14, 20, 14, 14, 14, 14, 16, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    # Freeze header row
    ws.freeze_panes = "A5"

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Build filename
    safe_period = period.replace(" ", "_")
    filename = f"Payments_{safe_period}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/agreements")
def export_agreements_excel(
    status: str = Query("", description="Filter by status: active, pending, expired, or empty for all"),
    search: str = Query("", description="Text search filter"),
    current_user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    """Export filtered agreements to an Excel (.xlsx) file. Admin only."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Export is restricted to administrators")

    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    cursor = db.cursor()

    # Query agreements with analysis data (same join as list_agreements)
    query = """
        SELECT a.*, aa.company_name, aa.agreement_title, aa.effective_date, aa.expiry_date,
               aa.payment_plans, aa.currency
        FROM agreements a
        LEFT JOIN agreement_analysis aa ON a.id = aa.agreement_id
        WHERE 1=1
    """
    params = []

    if status:
        query += " AND a.status = ?"
        params.append(status)

    if search:
        query += " AND (aa.company_name LIKE ? OR a.file_name LIKE ? OR CAST(a.id AS TEXT) LIKE ?)"
        search_term = f"%{search}%"
        params.extend([search_term, search_term, search_term])

    query += " ORDER BY a.uploaded_at DESC"
    rows = cursor.execute(query, params).fetchall()

    # Process agreements
    filtered = []
    for row in rows:
        r = dict(row)
        total_amount = 0
        paid_amount = 0
        plans_str = r.get("payment_plans")
        if plans_str:
            try:
                plans = json.loads(plans_str) if isinstance(plans_str, str) else plans_str
                if isinstance(plans, list):
                    for plan in plans:
                        if isinstance(plan, dict):
                            amt = plan.get("amount", 0) or 0
                            total_amount += amt
                            if (plan.get("status") or "").lower() == "paid":
                                paid_amount += amt
            except (json.JSONDecodeError, TypeError):
                pass

        filtered.append({
            "company_name": r.get("company_name") or r.get("file_name") or "Unknown",
            "agreement_title": r.get("agreement_title") or "—",
            "status": (r.get("status") or "pending").capitalize(),
            "effective_date": r.get("effective_date") or "",
            "expiry_date": r.get("expiry_date") or "",
            "total_amount": total_amount,
            "paid_amount": paid_amount,
            "currency": r.get("currency") or "₹",
        })

    # Build Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Agreements"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    total_font = Font(name="Calibri", bold=True, size=11)
    total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    active_font = Font(name="Calibri", color="059669")
    pending_font = Font(name="Calibri", color="D97706")
    expired_font = Font(name="Calibri", color="DC2626")
    currency_format = '#,##0.00'

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    status_label = f" — {status.capitalize()}" if status else ""
    title_cell.value = f"Agreements Report{status_label}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle row
    ws.merge_cells("A2:H2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    sub_cell.font = Font(name="Calibri", size=9, color="6B7280")
    ws.row_dimensions[2].height = 20

    # Headers (row 4)
    headers = ["Company", "Agreement Title", "Status", "Start Date", "End Date", "Total Amount", "Paid Amount", "Currency"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[4].height = 28

    # Data rows
    grand_total = 0
    grand_paid = 0

    for row_idx, a in enumerate(filtered, 5):
        ws.cell(row=row_idx, column=1, value=a["company_name"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=a["agreement_title"]).border = thin_border

        status_cell = ws.cell(row=row_idx, column=3, value=a["status"])
        status_cell.border = thin_border
        if a["status"] == "Active":
            status_cell.font = active_font
        elif a["status"] == "Pending":
            status_cell.font = pending_font
        elif a["status"] == "Expired":
            status_cell.font = expired_font

        # Dates
        for col, date_str in [(4, a["effective_date"]), (5, a["expiry_date"])]:
            try:
                date_val = datetime.strptime(date_str, "%Y-%m-%d")
                date_cell = ws.cell(row=row_idx, column=col, value=date_val)
                date_cell.number_format = "DD MMM YYYY"
            except (ValueError, TypeError):
                date_cell = ws.cell(row=row_idx, column=col, value=date_str or "—")
            date_cell.border = thin_border

        amount_cell = ws.cell(row=row_idx, column=6, value=a["total_amount"])
        amount_cell.number_format = currency_format
        amount_cell.border = thin_border

        paid_cell = ws.cell(row=row_idx, column=7, value=a["paid_amount"])
        paid_cell.number_format = currency_format
        paid_cell.font = Font(name="Calibri", color="059669") if a["paid_amount"] > 0 else Font(name="Calibri")
        paid_cell.border = thin_border

        ws.cell(row=row_idx, column=8, value=a["currency"]).border = thin_border

        grand_total += a["total_amount"]
        grand_paid += a["paid_amount"]

    # Total row
    total_row = len(filtered) + 5
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=1).border = thin_border
    for col in range(2, 6):
        ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=col).border = thin_border

    for col, val in [(6, grand_total), (7, grand_paid)]:
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.number_format = currency_format
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border

    ws.cell(row=total_row, column=8).fill = total_fill
    ws.cell(row=total_row, column=8).border = thin_border

    # Auto-size columns
    col_widths = [28, 28, 12, 16, 16, 16, 16, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    # Freeze header row
    ws.freeze_panes = "A5"

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Agreements_Report{'_' + status.capitalize() if status else ''}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
